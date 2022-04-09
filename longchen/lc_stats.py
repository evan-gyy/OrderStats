#!/usr/bin/env python
# -*- coding:utf-8 -*-
# author: evan-gyy
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, colors, Font
import traceback
import os
import gc

class LCStats:
    def __init__(self):
        self.file = ''
        self.find_file('.xlsx', '.')
        self.map = {}
        """
        self.data样例:
        '10': {
            'loc': [],
            'sum': 3,
            'red': 1,
            'order': {
                '601': {
                    '蛋': 1
                },
                '301': {
                    '蛋': 2
                }
            }
        },
        """
        self.data = {}
        self.wb = openpyxl.load_workbook(self.file)
        self.ws = self.wb.worksheets[0]

    def find_file(self, type, path):
        file_list = []
        for f in os.listdir(path):
            if type in f and 'res-' not in f:
                file_list.append(f)
        if len(file_list) > 1:
            print("检测到以下excel文件：")
            for f in file_list:
                print("{}：{}".format(file_list.index(f), f))
            while True:
                try:
                    self.file = file_list[int((input("请输入文件序号：")))]
                    break
                except:
                    print("发生错误：请正确输入文件前的序号（0-n）")
        elif len(file_list) == 1:
            self.file = file_list[0]
        else:
            input("请核对目录下是否有excel文件")
            exit()

    def get_data(self, sheet):
        df = pd.read_excel(self.file, sheet_name=sheet)
        for index, row in df.iterrows():
            if pd.isnull(row['跟团号']):
                break
            lou = int(row['楼号'])
            nong = int(row['弄号'])
            key = str(lou) if nong != 719 else '719-' + str(lou)
            room = str(int(row['房间号']))
            good = str(row['物资'])
            n = int(row['数量'])
            red = 1 if row['是否封控'] != '未' else 0
            if key not in self.data:
                self.data[key] = {
                    'loc': [],
                    'sum': 0,
                    'red': 0,
                    'order': {}
                }
            self.data[key]['sum'] += n
            self.data[key]['red'] = red
            if room not in self.data[key]['order']:
                self.data[key]['order'][room] = {}
            if good not in self.data[key]['order'][room]:
                self.data[key]['order'][room][good] = 0
            self.data[key]['order'][room][good] += n
        # print(self.data)

    def to_map(self):
        total = 0
        for i in range(1, self.ws.max_row + 1):
            sum = 0
            for j in range(2, self.ws.max_column + 1):
                cell = self.ws.cell(i, j).value
                # print(cell, type(cell))
                if not cell:
                    continue
                cell = str(cell)
                if cell not in self.data:
                    continue
                d = self.data[cell]
                sum += d['sum']
                if d['red']:
                    self.ws.cell(i, j).fill = PatternFill("solid", fgColor="FF0000")
                    self.ws.cell(i + 1, j).fill = PatternFill("solid", fgColor="FF0000")
                    self.ws.cell(i, j).font = Font('Times New Roman', bold=True, color="FFFFFF")
                    self.ws.cell(i + 1, j).font = Font('Times New Roman', bold=True, color="FFFFFF")
                else:
                    self.ws.cell(i, j).fill = PatternFill("solid", fgColor="FFC000")
                    self.ws.cell(i + 1, j).fill = PatternFill("solid", fgColor="FFC000")

                orders = []
                for room, order in d['order'].items():
                    for good, num in order.items():
                        info = room + good + str(num)
                        orders.append(info)
                self.ws.cell(i + 1, j).value = '\n'.join(orders)

            if sum:
                total += sum
                self.ws.cell(i + 1, 1).value = sum

        self.ws.cell(19, 1).value = total

    def run(self):
        sheets = self.wb.worksheets
        self.get_data(sheets[1].title)
        self.to_map()
        self.wb.save('res-' + self.file)
        del self.wb, self.ws
        gc.collect()

if __name__ == '__main__':
    try:
        lc = LCStats()
        lc.run()
    except:
        traceback.print_exc()
        input()